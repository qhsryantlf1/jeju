import base64
import io
import re
import urllib.parse
import urllib.request
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.styles.colors import COLOR_INDEX

WEEKDAYS = {"일", "월", "화", "수", "목", "금", "토"}
DEFAULT_SHEET_ID = "1iF35MTE-aRmYOaAWhUpTypASiI030VJV6rCGbcATF0I"
DEFAULT_COLOR = "#000000"
STUDENT_TAB = "학생정보"
SCHOOL_TAB = "학교 정보"
THEME_COLORS = {
    0: "#000000",
    1: "#000000",
    2: "#ff0000",
    3: "#00ff00",
    4: "#0000ff",
    5: "#ffff00",
    6: "#ff00ff",
    7: "#00ffff",
    9: "#ff9900",
    10: "#666666",
}


def sheet_xlsx_url(sheet_id: str, tab_name: str) -> str:
    query = urllib.parse.urlencode({"format": "xlsx", "sheet": tab_name})
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?{query}"


def fetch_sheet_xlsx(sheet_id: str, tab_name: str) -> bytes:
    url = sheet_xlsx_url(sheet_id, tab_name)
    request = urllib.request.Request(url, headers={"User-Agent": "jeju-tv-bulletin/1.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        return response.read()


def font_color_hex(font) -> str:
    if not font or not font.color:
        return DEFAULT_COLOR

    color = font.color
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:].lower()
        if rgb in {"ffffff", "000000"}:
            return DEFAULT_COLOR
        return f"#{rgb}"

    if color.type == "theme":
        return THEME_COLORS.get(color.theme, DEFAULT_COLOR)

    return DEFAULT_COLOR


def parse_styled_lines(cell) -> list[dict]:
    lines = []
    default_color = font_color_hex(cell.font)

    if isinstance(cell.value, CellRichText):
        for block in cell.value:
            text = str(block)
            color = font_color_hex(block.font) or default_color
            for part in text.split("\n"):
                part = part.strip()
                if part:
                    lines.append({"text": part, "color": color})
        return lines

    if cell.value in (None, ""):
        return lines

    text = str(cell.value).strip()
    color = default_color
    for part in text.split("\n"):
        part = part.strip()
        if part:
            lines.append({"text": part, "color": color})
    return lines


def expand_slash_styled(event_lines, dept_lines, guide_lines):
    out_events, out_event_colors = [], []
    out_depts, out_dept_colors = [], []
    out_guides, out_guide_colors = [], []

    if not event_lines:
        for guide in guide_lines:
            if not guide["text"]:
                continue
            out_events.append("")
            out_event_colors.append(DEFAULT_COLOR)
            out_depts.append("")
            out_dept_colors.append(DEFAULT_COLOR)
            out_guides.append(guide["text"])
            out_guide_colors.append(guide["color"])
        return (
            out_events,
            out_event_colors,
            out_depts,
            out_dept_colors,
            out_guides,
            out_guide_colors,
        )

    for idx, event in enumerate(event_lines):
        dept = dept_lines[idx] if idx < len(dept_lines) else (dept_lines[0] if len(dept_lines) == 1 else {"text": "", "color": DEFAULT_COLOR})
        guide = guide_lines[idx] if idx < len(guide_lines) else (guide_lines[0] if len(guide_lines) == 1 else {"text": "", "color": DEFAULT_COLOR})

        parts = [p.strip() for p in re.split(r"\s+/\s+", event["text"]) if p.strip()]
        split_events = parts if len(parts) > 1 else [event["text"]]

        for part in split_events:
            out_events.append(part)
            out_event_colors.append(event["color"])
            out_depts.append(dept["text"])
            out_dept_colors.append(dept["color"])
            out_guides.append(guide["text"])
            out_guide_colors.append(guide["color"])

    return (
        out_events,
        out_event_colors,
        out_depts,
        out_dept_colors,
        out_guides,
        out_guide_colors,
    )


def parse_title_meta(ws) -> dict:
    for row in range(1, 6):
        title = str(ws.cell(row, 1).value or "").strip()
        if not title:
            continue
        match = re.search(r"(\d{4})학년도\s*(\d{1,2})월", title)
        if match:
            return {
                "title": title,
                "year": int(match.group(1)),
                "month": int(match.group(2)),
            }
        match = re.search(r"(\d{1,2})월", title)
        if match:
            return {"title": title, "month": int(match.group(1))}
    return {}


def normalize_day(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    return None


def parse_sheet_workbook(ws) -> tuple[list, dict]:
    meta = parse_title_meta(ws)
    days = []

    for row in range(1, ws.max_row + 1):
        day = normalize_day(ws.cell(row, 1).value)
        weekday = str(ws.cell(row, 2).value or "").strip()
        if day is None or weekday not in WEEKDAYS:
            continue

        event_lines = parse_styled_lines(ws.cell(row, 3))
        dept_lines = parse_styled_lines(ws.cell(row, 4))
        guide_lines = parse_styled_lines(ws.cell(row, 5))

        (
            events,
            event_colors,
            departments,
            department_colors,
            life_guides,
            life_guide_colors,
        ) = expand_slash_styled(event_lines, dept_lines, guide_lines)

        days.append({
            "day": day,
            "weekday": weekday,
            "events": events,
            "eventColors": event_colors,
            "departments": departments,
            "departmentColors": department_colors,
            "lifeGuides": life_guides,
            "lifeGuideColors": life_guide_colors,
            "dayColor": font_color_hex(ws.cell(row, 1).font),
            "weekdayColor": font_color_hex(ws.cell(row, 2).font),
        })

    days.sort(key=lambda item: item["day"])
    return days, meta


def color_to_hex(color) -> str | None:
    if not color:
        return None

    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:].lower()
        if rgb in {"ffffff", "000000"}:
            return None
        return f"#{rgb}"

    if color.type == "theme":
        return THEME_COLORS.get(color.theme)

    if color.type == "indexed" and color.indexed is not None:
        if color.indexed in (64, 65):
            return None
        indexed = COLOR_INDEX.get(color.indexed)
        if indexed:
            rgb = indexed[-6:].lower()
            if rgb in {"ffffff", "000000"}:
                return None
            return f"#{rgb}"

    return None


def fill_color_hex(cell) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    return color_to_hex(fill.fgColor)


def format_cell_value(value) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"

    if isinstance(value, date):
        return f"{value.month}-{value.day}"

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"

    return str(value).strip()


def worksheet_bounds(ws) -> tuple[int, int, int, int]:
    min_row = ws.max_row
    min_col = ws.max_column
    max_row = 0
    max_col = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            min_row = min(min_row, cell.row)
            min_col = min(min_col, cell.column)
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)

    if max_row == 0:
        return 1, 1, 1, 1

    return min_row, max_row, min_col, max_col


def build_merge_maps(ws) -> tuple[dict, set]:
    merge_info = {}
    skip = set()

    for merged in ws.merged_cells.ranges:
        min_row, min_col = merged.min_row, merged.min_col
        rowspan = merged.max_row - merged.min_row + 1
        colspan = merged.max_col - merged.min_col + 1
        merge_info[(min_row, min_col)] = {"rowspan": rowspan, "colspan": colspan}

        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if row != min_row or col != min_col:
                    skip.add((row, col))

    return merge_info, skip


def parse_table_worksheet(ws) -> dict:
    min_row, max_row, min_col, max_col = worksheet_bounds(ws)
    merge_info, skip = build_merge_maps(ws)
    rows = []

    for row_idx in range(min_row, max_row + 1):
        cells = []
        for col_idx in range(min_col, max_col + 1):
            if (row_idx, col_idx) in skip:
                continue

            cell = ws.cell(row_idx, col_idx)
            value = format_cell_value(cell.value)
            merge = merge_info.get((row_idx, col_idx))

            if not value and not merge:
                continue

            item = {
                "value": value,
                "color": font_color_hex(cell.font),
                "bg": fill_color_hex(cell),
                "bold": bool(cell.font and cell.font.bold),
            }

            if merge:
                if merge["rowspan"] > 1:
                    item["rowspan"] = merge["rowspan"]
                if merge["colspan"] > 1:
                    item["colspan"] = merge["colspan"]

            cells.append(item)

        if cells:
            rows.append({"cells": cells})

    title = ""
    if rows and rows[0]["cells"]:
        title = rows[0]["cells"][0].get("value", "")

    return {"title": title, "rows": rows}


def extract_worksheet_banner_image(ws) -> dict | None:
    images = getattr(ws, "_images", None) or []
    if not images:
        return None

    img = images[0]
    try:
        raw = img._data()
    except Exception:
        return None

    fmt = (getattr(img, "format", None) or "png").lower()
    if fmt == "jpg":
        fmt = "jpeg"
    mime = f"image/{fmt}"

    aspect = 114 / 1096
    anchor = getattr(img, "anchor", None)
    if anchor is not None and getattr(anchor, "ext", None) is not None:
        ext = anchor.ext
        if ext.cx and ext.cy:
            aspect = ext.cy / ext.cx

    return {
        "dataUrl": f"data:{mime};base64,{base64.b64encode(raw).decode('ascii')}",
        "aspect": aspect,
    }


def parse_status_sheets(sheet_id: str) -> dict:
    student_wb = load_workbook(
        io.BytesIO(fetch_sheet_xlsx(sheet_id, STUDENT_TAB)),
        data_only=False,
        rich_text=False,
    )
    school_wb = load_workbook(
        io.BytesIO(fetch_sheet_xlsx(sheet_id, SCHOOL_TAB)),
        data_only=True,
        rich_text=False,
    )

    student_ws = student_wb[STUDENT_TAB] if STUDENT_TAB in student_wb.sheetnames else student_wb.worksheets[0]
    school_ws = school_wb[SCHOOL_TAB] if SCHOOL_TAB in school_wb.sheetnames else school_wb.worksheets[0]

    student_data = parse_table_worksheet(student_ws)
    title_image = extract_worksheet_banner_image(student_ws)
    if title_image:
        student_data["titleImage"] = title_image

    return {
        "student": student_data,
        "school": parse_table_worksheet(school_ws),
        "meta": {
            "source": "google-sheet",
            "sheetId": sheet_id,
            "studentTab": STUDENT_TAB,
            "schoolTab": SCHOOL_TAB,
        },
    }


def parse_google_sheet(sheet_id: str, tab_name: str) -> dict:
    workbook = load_workbook(
        io.BytesIO(fetch_sheet_xlsx(sheet_id, tab_name)),
        data_only=False,
        rich_text=True,
    )
    worksheet = workbook[tab_name] if tab_name in workbook.sheetnames else workbook.worksheets[0]
    data, meta = parse_sheet_workbook(worksheet)
    if not data:
        raise ValueError("시트에서 일정 데이터를 찾지 못했습니다.")

    return {
        "data": data,
        "meta": {
            "source": "google-sheet",
            "sheetId": sheet_id,
            "tab": tab_name,
            **meta,
        },
    }
